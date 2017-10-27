import openpyxl
from flask import send_from_directory
from flask_jwt_extended import get_jwt_identity, jwt_required
from flask_restful_swagger_2 import Resource, swagger

from db.models.account import AdminModel, StudentModel

from . import goingout_doc


class Goingout(Resource):
    @swagger.doc(goingout_doc.GOINGOUT_GET)
    @jwt_required
    def get(self):
        """
        외출신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=get_jwt_identity())

        if not admin:
            return '', 403

        wb = openpyxl.load_workbook('명렬표.xlsx')
        ws = wb.active

        for row in map(str, range(3, 68)):
            for column1, column2, column3 in zip(['B', 'F', 'J', 'N'], ['D', 'H', 'L', 'P'], ['E', 'I', 'M', 'Q']):
                if ws[column1+row].value == '학번':
                    continue

                number = int(ws[column1 + row].value) if ws[column1 + row].value else None
                if not number:
                    continue

                student = StudentModel.objects(number=number).first()
                if not student or not student.goingout_apply:
                    continue

                sat = '토요 외출' if student.goingout_apply.on_saturday else ''
                sun = '일요 외출' if student.goingout_apply.on_sunday else ''
                ws[column2+row] = sat
                ws[column3+row] = sun

        wb.save('명렬표.xlsx')

        return send_from_directory('.', '명렬표.xlsx')
